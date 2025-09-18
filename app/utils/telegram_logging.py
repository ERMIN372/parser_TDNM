from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

from aiogram import Bot
from aiogram.types import InputFile

from .logging import log_event, mark_response


def _resolve_document_path(document: Any) -> Path | None:
    if isinstance(document, InputFile):
        source = getattr(document, "path", None) or getattr(document, "path_or_bytesio", None)
        if isinstance(source, (str, Path)):
            return Path(source)
    if isinstance(document, (str, Path)):
        return Path(document)
    return None


def _detect_rows(path: Path | None) -> int | None:
    if not path or not path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover - optional dependency missing
        return None
    try:
        wb = load_workbook(path, read_only=True)
        sheet = wb.active
        rows = max(0, (sheet.max_row or 0) - 1)
        wb.close()
        return rows
    except Exception:  # pragma: no cover - tolerate parsing errors
        return None


def _document_meta(result_message) -> Dict[str, Any]:
    if not result_message or not getattr(result_message, "document", None):
        return {}
    doc = result_message.document
    meta: Dict[str, Any] = {}
    if getattr(doc, "file_name", None):
        meta["filename"] = doc.file_name
    if getattr(doc, "file_size", None):
        meta["size_kb"] = round(doc.file_size / 1024, 1)
    return meta


class LoggedBot(Bot):
    async def send_message(self, chat_id, text, **kwargs):  # type: ignore[override]
        response = await super().send_message(chat_id, text, **kwargs)
        mark_response(reply_type="text", text=text)
        return response

    async def send_photo(self, chat_id, photo, **kwargs):  # type: ignore[override]
        caption = kwargs.get("caption")
        response = await super().send_photo(chat_id, photo, **kwargs)
        mark_response(reply_type="photo", text=caption)
        return response

    async def send_document(self, chat_id, document, **kwargs):  # type: ignore[override]
        caption = kwargs.get("caption")
        response = await super().send_document(chat_id, document, **kwargs)
        meta = _document_meta(response)
        if "rows" not in meta:
            meta_path = _resolve_document_path(document)
            rows = _detect_rows(meta_path)
            if rows is not None:
                meta["rows"] = rows
        log_event("file_sent", message=f"document sent {meta.get('filename', '')}", document_meta=meta)
        mark_response(reply_type="document", text=caption, document_meta=meta)
        return response

    async def send_invoice(self, chat_id, **kwargs):  # type: ignore[override]
        response = await super().send_invoice(chat_id, **kwargs)
        title = kwargs.get("title")
        mark_response(reply_type="invoice", text=title)
        return response
