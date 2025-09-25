"""Utilities for collecting diagnostics about generated XLSX reports."""

from __future__ import annotations

import hashlib
import stat
import warnings
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:  # pragma: no cover - optional dependency is part of runtime image
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - openpyxl may not be available in tests
    load_workbook = None  # type: ignore[assignment]


MAX_ZIP_ENTRIES = 20
MAX_SHEETS = 20
HASH_SIZE_LIMIT_MB = 50


@dataclass(slots=True)
class XlsxDiagnostics:
    """Snapshot with metadata describing the XLSX file."""

    path: str
    exists: bool
    size_bytes: int | None = None
    modified_ts: int | None = None
    permissions: str | None = None
    signature_hex: str | None = None
    sha256: str | None = None
    is_zip: bool | None = None
    zip_valid: bool | None = None
    zip_entry_count: int | None = None
    zip_entries: list[str] | None = None
    sheet_count: int | None = None
    sheet_names: list[str] | None = None
    errors: list[str] = field(default_factory=list)

    def add_error(self, message: str) -> None:
        message = message.strip()
        if message and message not in self.errors:
            self.errors.append(message)

    def to_event_payload(self) -> dict[str, Any]:
        payload: dict[str, Any] = {"path": self.path, "exists": self.exists}
        optional: dict[str, Any] = {
            "size_bytes": self.size_bytes,
            "modified_ts": self.modified_ts,
            "permissions": self.permissions,
            "signature_hex": self.signature_hex,
            "sha256": self.sha256,
            "is_zip": self.is_zip,
            "zip_valid": self.zip_valid,
            "zip_entry_count": self.zip_entry_count,
            "zip_entries": self.zip_entries,
            "sheet_count": self.sheet_count,
            "sheet_names": self.sheet_names,
        }
        for key, value in optional.items():
            if value is not None and value != []:
                payload[key] = value
        if self.errors:
            payload["errors"] = list(self.errors)
        return payload


def collect_xlsx_diagnostics(path: Path | str) -> XlsxDiagnostics:
    """Collect metadata and lightweight integrity information about XLSX file."""

    path_obj = Path(path)
    diagnostics = XlsxDiagnostics(path=str(path_obj), exists=path_obj.exists())

    if not diagnostics.exists:
        diagnostics.add_error("file_not_found")
        return diagnostics

    try:
        stat_result = path_obj.stat()
        diagnostics.size_bytes = stat_result.st_size
        diagnostics.modified_ts = int(stat_result.st_mtime)
        diagnostics.permissions = oct(stat.S_IMODE(stat_result.st_mode))
    except OSError as exc:
        diagnostics.add_error(f"stat:{exc}")

    # Read signature and optionally hash the file (bounded by size limit)
    try:
        file_size = diagnostics.size_bytes or path_obj.stat().st_size
        hash_limit_bytes = HASH_SIZE_LIMIT_MB * 1024 * 1024
        should_hash = file_size <= hash_limit_bytes

        with path_obj.open("rb") as src:
            head = src.read(16)
            diagnostics.signature_hex = head.hex() if head else None
            if should_hash:
                hasher = hashlib.sha256()
                if head:
                    hasher.update(head)
                for chunk in iter(lambda: src.read(65536), b""):
                    hasher.update(chunk)
                diagnostics.sha256 = hasher.hexdigest()
            else:
                diagnostics.add_error("sha256_skipped_large_file")
    except OSError as exc:
        diagnostics.add_error(f"read:{exc}")

    try:
        diagnostics.is_zip = zipfile.is_zipfile(path_obj)
    except OSError as exc:
        diagnostics.add_error(f"zipcheck:{exc}")

    if diagnostics.is_zip:
        try:
            with zipfile.ZipFile(path_obj) as archive:
                names = archive.namelist()
                diagnostics.zip_entry_count = len(names)
                diagnostics.zip_entries = names[:MAX_ZIP_ENTRIES]
                if len(names) > MAX_ZIP_ENTRIES:
                    diagnostics.zip_entries.append("…")
                diagnostics.zip_valid = True
        except zipfile.BadZipFile as exc:
            diagnostics.zip_valid = False
            diagnostics.add_error(f"bad_zip:{exc}")
        except Exception as exc:  # pragma: no cover - defensive logging
            diagnostics.zip_valid = False
            diagnostics.add_error(f"zip_error:{exc}")

    if load_workbook is None:
        diagnostics.add_error("openpyxl_unavailable")
        return diagnostics

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(path_obj, read_only=True, data_only=True)
        try:
            names = list(workbook.sheetnames)
        finally:
            workbook.close()
        diagnostics.sheet_count = len(names)
        diagnostics.sheet_names = names[:MAX_SHEETS]
        if len(names) > MAX_SHEETS:
            diagnostics.sheet_names.append("…")
    except Exception as exc:  # pragma: no cover - workbook inspection best effort
        diagnostics.add_error(f"openpyxl:{exc}")

    return diagnostics


__all__ = ["collect_xlsx_diagnostics", "XlsxDiagnostics"]
